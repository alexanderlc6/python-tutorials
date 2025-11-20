import openpyxl as pxl

excel = pxl.load_workbook('../../resources/namelist.xlsx')
sheet = excel['Sheet1']

rows = sheet.max_row
cols = sheet.max_column

#Normally read data
# for i in range(1, rows+1):
#     for j in range(1, cols+1):
#         cellVal = sheet.cell(i,j).value
#         print('%s\t'%(cellVal),end='')
#     print('')

#Find all students whose grade greater than 90
#Key:class name, value: grade list
result = dict()
for i in range(2, rows+1):
    classroom = sheet.cell(i, 1).value
    name = sheet.cell(i, 2).value
    score = int(sheet.cell(i, 3).value)

    if score >= 90:
        if classroom not in result:
            result[classroom] = list()
            result[classroom].append(score)

output = open('../../resources/scores.txt', 'w')
for classroom in result:
    for score in result[classroom]:
        output.write('%s\t%s\n'%(classroom, str(score)))

excel.close()
output.close()